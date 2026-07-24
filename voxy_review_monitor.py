import argparse
import hashlib
import json
import os
import re
import smtplib
import sys
from dataclasses import dataclass
from datetime import date, datetime, timezone
from email.message import EmailMessage
from pathlib import Path
from typing import Iterable
from zoneinfo import ZoneInfo

try:
    from bs4 import BeautifulSoup
    from dotenv import load_dotenv
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from playwright.sync_api import sync_playwright
    import gspread
    import requests
except ImportError as exc:
    print(f"Module manquant: {exc.name}")
    print("Installe les dependances avec: pip install -r requirements.txt")
    raise


@dataclass
class Product:
    name: str
    url: str
    emails: list[str]
    threshold: float
    platform: str
    language: str
    country: str = ""
    owner: str = ""
    priority: str = "Medium"
    alert_type: str = "Both"
    paused: bool = False
    score_threshold: float = 3.0
    platform_account: str = ""


@dataclass
class Review:
    rating: float
    text: str
    author: str = ""
    date: str = ""
    source: str = ""

    @property
    def fingerprint(self) -> str:
        clean_text = normalize_text(self.text)[:700]
        raw = f"{self.rating}|{self.author}|{self.date}|{clean_text}"
        return hashlib.sha256(raw.encode("utf-8", errors="ignore")).hexdigest()


THEME_KEYWORDS = {
    "Guide / staff": ["guide", "staff", "host", "driver", "employee", "rude", "friendly", "professional", "helpful"],
    "Organization": ["organization", "organised", "organized", "waiting", "wait", "late", "delay", "queue", "meeting point"],
    "Value for money": ["price", "expensive", "value", "money", "worth", "overpriced", "cost"],
    "Communication": ["communication", "instructions", "information", "email", "message", "unclear", "confusing"],
    "Booking / access": ["ticket", "booking", "reservation", "entry", "access", "cancel", "refund", "voucher"],
    "Experience quality": ["experience", "tour", "activity", "boring", "amazing", "interesting", "disappointing"],
}


IMPROVEMENT_LIBRARY = {
    "Guide / staff": "Review guide/staff briefing, tone, and service consistency.",
    "Organization": "Reduce waiting time, clarify the meeting process, and check day-of operations.",
    "Value for money": "Check perceived value: inclusions, price positioning, and expectation setting.",
    "Communication": "Improve pre-arrival instructions, confirmation messages, and on-page clarity.",
    "Booking / access": "Audit booking, ticketing, refund, and entry instructions end to end.",
    "Experience quality": "Compare the advertised promise with the actual customer experience.",
}


DETAILED_IMPROVEMENTS = {
    "Guide / staff": [
        "Audit the guide briefing: key facts, timing, tone, and expected storytelling flow.",
        "Review low-rated guide mentions and coach recurring issues such as pace, clarity, or engagement.",
        "Create a short quality checklist for guides before each departure.",
    ],
    "Organization": [
        "Check meeting-point instructions against the real customer journey on site.",
        "Reduce waiting time and make late-start escalation rules explicit.",
        "Add clearer pre-tour timing and arrival instructions to confirmation messages.",
    ],
    "Value for money": [
        "Compare the product promise, inclusions, and price with the actual delivered experience.",
        "Clarify what is included and what is not included on the product page.",
        "Use recent critical reviews to identify whether customers expected more access, time, or guidance.",
    ],
    "Communication": [
        "Rewrite the pre-arrival instructions in simpler, step-by-step language.",
        "Add one clear contact/escalation instruction for customers who cannot find the guide.",
        "Check whether voucher, meeting point, and entry instructions say the same thing everywhere.",
    ],
    "Booking / access": [
        "Audit the booking-to-entry flow: voucher, reserved entry, ticket scan, and staff handoff.",
        "Clarify entry restrictions, security lines, and what 'reserved' or 'priority' means.",
        "Review cancellation/refund complaints and standardize the response script.",
    ],
    "Experience quality": [
        "Compare the advertised highlights with the parts customers actually mention in reviews.",
        "Identify the weakest moments in the tour flow and redesign transitions or explanations.",
        "Add a post-tour quality check for repeated disappointment signals.",
    ],
    "General satisfaction": [
        "Read the latest low-rated reviews manually and group them by operational root cause.",
        "Prioritize fixes that appear across several reviews or across several products.",
        "Track whether the same complaint returns after the operational fix is made.",
    ],
}


def split_emails(value: str) -> list[str]:
    return [email.strip() for email in re.split(r"[;,]", value or "") if email.strip()]


def normalize_text(value: str) -> str:
    return " ".join(str(value or "").replace("\xa0", " ").split())


def truthy(value: str) -> bool:
    return str(value or "").strip().lower() in {"oui", "yes", "true", "1", "x"}


def optional_float(value, default: float) -> float:
    text = str(value or "").strip().replace(",", ".")
    if not text:
        return default
    try:
        return float(text)
    except ValueError:
        return default


def decode_jsonish_text(value: str) -> str:
    text = str(value or "")
    text = text.replace("\\u002F", "/")
    text = text.replace("\\/", "/")
    try:
        text = bytes(text, "utf-8").decode("unicode_escape")
    except Exception:
        pass
    return text


def extract_json_field_text(value: str, field: str) -> str:
    text = decode_jsonish_text(value)
    match = re.search(rf'"{re.escape(field)}"\s*:\s*"((?:\\.|[^"\\])*)"', text, re.I)
    if not match:
        return ""
    return normalize_text(decode_jsonish_text(match.group(1)))


def clean_review_text_for_email(value: str, max_chars: int = 420) -> str:
    text = decode_jsonish_text(value)
    message = extract_json_field_text(text, "message")
    if message:
        text = message
    else:
        text = html_to_text(text)

    # Remove obvious page/config fragments when extraction captured surrounding JSON.
    text = re.sub(r'\{?("__entity"|type|content|media|review|author|rating|url)"?\s*:\s*', " ", text)
    text = re.sub(r'https?://\S+', " ", text)
    text = normalize_text(text)

    if len(text) > max_chars:
        text = text[:max_chars].rsplit(" ", 1)[0] + "..."
    return text


def clean_author_for_email(review: Review) -> str:
    author = review.author or extract_json_field_text(review.text, "title") or extract_json_field_text(review.text, "shortTitle")
    return normalize_text(author) or "not detected"


def clean_date_for_email(review: Review) -> str:
    date_value = review.date or extract_json_field_text(review.text, "subtitle") or extract_json_field_text(review.text, "shortSubtitle")
    return normalize_text(date_value) or "not detected"


def trend_with_arrow(summary: dict | None) -> str:
    if not summary:
        return "No history"
    trend = summary.get("trend", "No history")
    if trend.startswith("Improving"):
        return f"🟢 ↑ {trend}"
    if trend.startswith("Declining"):
        return f"🔴 ↓ {trend}"
    if trend.startswith("Stable"):
        return f"⚪ → {trend}"
    return f"⚪ → {trend}"


def product_health(summary: dict | None) -> str:
    if not summary:
        return "Needs attention"
    score = summary.get("global_score")
    critical = summary.get("critical_review_count", 0)
    low = summary.get("low_review_count", 0)
    trend = summary.get("trend", "")
    if summary.get("status") == "ERROR":
        return "Technical check needed"
    if score is not None and score < summary.get("score_threshold", 3.0):
        return "Bad"
    if critical > 0 or low >= 3 or trend.startswith("Declining"):
        return "Watch"
    return "Good"


def product_health_with_icon(summary: dict | None) -> str:
    health = product_health(summary)
    if health == "Good":
        return f"🟢 {health}"
    if health == "Bad":
        return f"🔴 {health}"
    if health == "Technical check needed":
        return f"🟠 {health}"
    return f"🟡 {health}"


def count_occurrences(text: str, keyword: str) -> int:
    return len(re.findall(re.escape(keyword), text, re.I))


def source_summary(reviews: list[Review]) -> str:
    if not reviews:
        return "No review source detected."
    counts: dict[str, int] = {}
    for review in reviews:
        source = review.source or "not specified"
        counts[source] = counts.get(source, 0) + 1
    return "\n".join(f"{source}: {count}" for source, count in sorted(counts.items(), key=lambda item: item[1], reverse=True))


def representative_examples(reviews: list[Review], limit: int = 3) -> list[str]:
    examples = []
    for review in sorted(reviews, key=lambda item: item.rating):
        text = clean_review_text_for_email(review.text, max_chars=220)
        if not text:
            continue
        examples.append(f"{review.rating}/5 - {text}")
        if len(examples) >= limit:
            break
    return examples


def analyze_improvement_opportunities(reviews: list[Review]) -> tuple[list[str], list[str]]:
    if not reviews:
        return (
            ["No current review text was available for detailed improvement analysis."],
            ["No review sources were available beyond page-level extraction."],
        )

    full_text = " ".join(clean_review_text_for_email(review.text, max_chars=1000).lower() for review in reviews if review.text)
    low_reviews = [review for review in reviews if review.rating <= 4]
    theme_scores = []
    for theme, keywords in THEME_KEYWORDS.items():
        count = sum(count_occurrences(full_text, keyword) for keyword in keywords)
        if count:
            theme_scores.append((theme, count))
    theme_scores.sort(key=lambda item: item[1], reverse=True)

    if not theme_scores:
        theme_scores = [("General satisfaction", 1)]

    opportunities = []
    for theme, count in theme_scores[:4]:
        examples = representative_examples(
            [review for review in low_reviews if any(keyword in clean_review_text_for_email(review.text).lower() for keyword in THEME_KEYWORDS.get(theme, []))]
            or low_reviews
            or reviews,
            limit=2,
        )
        actions = DETAILED_IMPROVEMENTS.get(theme, DETAILED_IMPROVEMENTS["General satisfaction"])
        opportunity = [
            f"{theme} ({count} signal{'s' if count != 1 else ''})",
            "Recommended actions:",
            *[f"- {action}" for action in actions[:3]],
        ]
        if examples:
            opportunity.extend(["Evidence examples:", *[f"- {example}" for example in examples]])
        opportunities.append("\n".join(opportunity))

    sources = [
        "Sources encountered:",
        source_summary(reviews),
        "",
        "Extraction note:",
        "Voxy uses structured page data, embedded rating patterns, visible review blocks, and page metadata when available. Some platforms expose partial JSON rather than clean review text, so Voxy cleans the text before writing the dashboard/email.",
    ]
    return opportunities, ["\n".join(sources)]


def build_global_synthesis(summaries: list[dict]) -> dict:
    product_count = len(summaries)
    total_reviews = sum(item["review_count"] for item in summaries)
    low_reviews = sum(item["low_review_count"] for item in summaries)
    critical_reviews = sum(item["critical_review_count"] for item in summaries)
    scored = [item["global_score"] for item in summaries if item["global_score"] is not None]
    avg_score = round(sum(scored) / len(scored), 2) if scored else None
    error_products = [item for item in summaries if item.get("status") == "ERROR"]

    theme_counts: dict[str, int] = {}
    for item in summaries:
        for theme in item.get("themes", []):
            theme_counts[theme] = theme_counts.get(theme, 0) + 1
    top_themes = [theme for theme, _ in sorted(theme_counts.items(), key=lambda item: item[1], reverse=True)[:5]]

    priority_actions = []
    for theme in top_themes or ["General satisfaction"]:
        actions = DETAILED_IMPROVEMENTS.get(theme, DETAILED_IMPROVEMENTS["General satisfaction"])
        priority_actions.append(f"{theme}: {actions[0]}")
    if error_products:
        priority_actions.append(f"Technical follow-up: {len(error_products)} product(s) had extraction errors or platform limitations.")

    sources = []
    for item in summaries:
        sources.append(f"{item['product']}: {item.get('sources_encountered', 'No source detected.')}")

    return {
        "product_count": product_count,
        "total_reviews": total_reviews,
        "average_score": avg_score,
        "low_reviews": low_reviews,
        "critical_reviews": critical_reviews,
        "top_themes": top_themes,
        "priority_actions": priority_actions,
        "sources": sources,
    }


def html_to_text(value: str) -> str:
    return normalize_text(BeautifulSoup(str(value or ""), "html.parser").get_text(" ", strip=True))


def parse_review_date(value: str) -> date | None:
    text = normalize_text(value)
    if not text:
        return None

    iso_match = re.search(r"\d{4}-\d{2}-\d{2}", text)
    if iso_match:
        try:
            return datetime.strptime(iso_match.group(0), "%Y-%m-%d").date()
        except ValueError:
            pass

    common_formats = [
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d-%m-%Y",
        "%m-%d-%Y",
        "%d %B %Y",
        "%B %d, %Y",
        "%d %b %Y",
        "%b %d, %Y",
    ]
    for fmt in common_formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def is_past_dated_review(review: Review, today: date) -> bool:
    parsed = parse_review_date(review.date)
    return parsed is not None and parsed < today


def load_products(xlsx_path: Path) -> list[Product]:
    workbook = load_workbook(xlsx_path, data_only=True)
    sheet = workbook["Produits"]
    headers = {
        str(cell.value).strip().lower(): index
        for index, cell in enumerate(sheet[1], start=1)
        if cell.value
    }

    active_header = "active" if "active" in headers else "actif"
    name_header = "product name" if "product name" in headers else "nom produit"
    url_header = "url to monitor" if "url to monitor" in headers else ("url a surveiller" if "url a surveiller" in headers else "url getyourguide")
    emails_header = "alert emails" if "alert emails" in headers else "emails alerte"
    threshold_header = "star threshold" if "star threshold" in headers else "seuil etoiles"
    platform_header = "platform" if "platform" in headers else "plateforme"
    language_header = "language" if "language" in headers else "langue"
    country_header = "country" if "country" in headers else "pays"
    owner_header = "owner" if "owner" in headers else "responsable"
    priority_header = "priority" if "priority" in headers else "priorite"
    alert_type_header = "alert type" if "alert type" in headers else "type alerte"
    paused_header = "paused" if "paused" in headers else "pause"
    score_threshold_header = "score alert threshold" if "score alert threshold" in headers else "seuil score"
    platform_account_header = "platform account" if "platform account" in headers else "compte plateforme"

    required = [active_header, name_header, url_header, emails_header]
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError(f"Missing columns in the Produits sheet: {', '.join(missing)}")

    products: list[Product] = []
    for row in range(2, sheet.max_row + 1):
        active = str(sheet.cell(row, headers[active_header]).value or "").strip().lower()
        name = str(sheet.cell(row, headers[name_header]).value or "").strip()
        url = str(sheet.cell(row, headers[url_header]).value or "").strip()
        emails = split_emails(str(sheet.cell(row, headers[emails_header]).value or ""))
        threshold_cell = headers.get(threshold_header)
        platform_cell = headers.get(platform_header)
        language_cell = headers.get(language_header)
        country_cell = headers.get(country_header)
        owner_cell = headers.get(owner_header)
        priority_cell = headers.get(priority_header)
        alert_type_cell = headers.get(alert_type_header)
        paused_cell = headers.get(paused_header)
        score_threshold_cell = headers.get(score_threshold_header)
        platform_account_cell = headers.get(platform_account_header)
        threshold = sheet.cell(row, threshold_cell).value if threshold_cell else 4
        platform = str(sheet.cell(row, platform_cell).value or "auto").strip().lower() if platform_cell else "auto"
        language = str(sheet.cell(row, language_cell).value or "en").strip() if language_cell else "en"
        country = str(sheet.cell(row, country_cell).value or "").strip() if country_cell else ""
        owner = str(sheet.cell(row, owner_cell).value or "").strip() if owner_cell else ""
        priority = str(sheet.cell(row, priority_cell).value or "Medium").strip() if priority_cell else "Medium"
        alert_type = str(sheet.cell(row, alert_type_cell).value or "Both").strip() if alert_type_cell else "Both"
        paused = truthy(str(sheet.cell(row, paused_cell).value or "")) if paused_cell else False
        score_threshold = optional_float(sheet.cell(row, score_threshold_cell).value if score_threshold_cell else "", 3.0)
        platform_account = str(sheet.cell(row, platform_account_cell).value or "").strip() if platform_account_cell else ""

        if active not in {"oui", "yes", "true", "1", "x"} or paused:
            continue
        if not name and not url:
            continue
        if not name or not url or not emails:
            print(f"Row {row} skipped: product name, URL, or alert emails are missing.")
            continue

        products.append(Product(
            name=name,
            url=url,
            emails=emails,
            threshold=optional_float(threshold, 4.0),
            platform=platform,
            language=language,
            country=country,
            owner=owner,
            priority=priority,
            alert_type=alert_type,
            paused=paused,
            score_threshold=score_threshold,
            platform_account=platform_account,
        ))
    return products


def products_from_rows(rows: list[list]) -> list[Product]:
    if not rows:
        return []
    headers = {
        str(value).strip().lower(): index
        for index, value in enumerate(rows[0])
        if value
    }

    active_col = headers.get("active", headers.get("actif"))
    name_col = headers.get("product name", headers.get("nom produit"))
    url_col = headers.get("url to monitor", headers.get("url a surveiller", headers.get("url getyourguide")))
    emails_col = headers.get("alert emails", headers.get("emails alerte"))
    threshold_col = headers.get("star threshold", headers.get("seuil etoiles"))
    platform_col = headers.get("platform", headers.get("plateforme"))
    language_col = headers.get("language", headers.get("langue"))
    country_col = headers.get("country", headers.get("pays"))
    owner_col = headers.get("owner", headers.get("responsable"))
    priority_col = headers.get("priority", headers.get("priorite"))
    alert_type_col = headers.get("alert type", headers.get("type alerte"))
    paused_col = headers.get("paused", headers.get("pause"))
    score_threshold_col = headers.get("score alert threshold", headers.get("seuil score"))
    platform_account_col = headers.get("platform account", headers.get("compte plateforme"))

    if None in {active_col, name_col, url_col, emails_col}:
        raise ValueError("Missing required columns in Google Sheet. Expected: Active, Product name, URL to monitor, Alert emails.")

    products: list[Product] = []
    for row_number, row in enumerate(rows[1:], start=2):
        def cell(index, default=""):
            return row[index] if index is not None and index < len(row) else default

        active = str(cell(active_col)).strip().lower()
        name = str(cell(name_col)).strip()
        url = str(cell(url_col)).strip()
        emails = split_emails(str(cell(emails_col)))
        threshold = cell(threshold_col, 4)
        platform = str(cell(platform_col, "auto")).strip().lower()
        language = str(cell(language_col, "en")).strip()
        country = str(cell(country_col, "")).strip()
        owner = str(cell(owner_col, "")).strip()
        priority = str(cell(priority_col, "Medium")).strip() or "Medium"
        alert_type = str(cell(alert_type_col, "Both")).strip() or "Both"
        paused = truthy(cell(paused_col, ""))
        score_threshold = optional_float(cell(score_threshold_col, ""), 3.0)
        platform_account = str(cell(platform_account_col, "")).strip()

        if active not in {"oui", "yes", "true", "1", "x"} or paused:
            continue
        if not name and not url:
            continue
        if not name or not url or not emails:
            print(f"Row {row_number} skipped: product name, URL, or alert emails are missing.")
            continue

        products.append(Product(
            name=name,
            url=url,
            emails=emails,
            threshold=optional_float(threshold, 4.0),
            platform=platform,
            language=language,
            country=country,
            owner=owner,
            priority=priority,
            alert_type=alert_type,
            paused=paused,
            score_threshold=score_threshold,
            platform_account=platform_account,
        ))
    return products


def google_client_from_env():
    service_account_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    if not service_account_json:
        return None
    credentials = json.loads(service_account_json)
    return gspread.service_account_from_dict(credentials)


def load_products_from_google_sheet(sheet_url: str) -> list[Product]:
    client = google_client_from_env()
    if client is None:
        raise RuntimeError("GOOGLE_SERVICE_ACCOUNT_JSON is required to read products directly from Google Sheets.")
    spreadsheet = client.open_by_url(sheet_url)
    worksheet = spreadsheet.worksheet("Produits")
    rows = worksheet.get_all_values()
    products = products_from_rows(rows)
    print(f"Loaded {len(products)} active product(s) directly from Google Sheets.")
    return products


def google_sheet_export_url(sheet_url: str) -> str:
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", sheet_url)
    if not match:
        raise ValueError("Invalid Google Sheets URL. Expected a URL containing /spreadsheets/d/{sheet_id}.")
    sheet_id = match.group(1)
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"


def download_google_sheet(sheet_url: str, destination: Path) -> Path:
    export_url = google_sheet_export_url(sheet_url)
    response = requests.get(export_url, timeout=60)
    response.raise_for_status()
    content_type = response.headers.get("content-type", "")
    if "text/html" in content_type.lower():
        raise RuntimeError(
            "Google returned an HTML page instead of an Excel file. "
            "Make sure the Google Sheet is shared so this PC can access it, or publish/export it with link access."
        )
    destination.write_bytes(response.content)
    return destination


def detect_platform(product: Product) -> str:
    if product.platform and product.platform != "auto":
        return product.platform
    url = product.url.lower()
    known = {
        "getyourguide": "getyourguide",
        "tripadvisor": "tripadvisor",
        "trustpilot": "trustpilot",
        "booking.com": "booking",
        "viator": "viator",
        "google": "google",
        "airbnb": "airbnb",
    }
    for needle, name in known.items():
        if needle in url:
            return name
    return "other"


def flatten_jsonld(item) -> Iterable[dict]:
    if isinstance(item, list):
        for child in item:
            yield from flatten_jsonld(child)
    elif isinstance(item, dict):
        yield item
        for value in item.values():
            if isinstance(value, (dict, list)):
                yield from flatten_jsonld(value)


def rating_from_value(value) -> float | None:
    if value is None:
        return None
    match = re.search(r"\d+(?:[.,]\d+)?", str(value))
    if not match:
        return None
    return float(match.group(0).replace(",", "."))


def first_text_value(obj: dict, keys: list[str]) -> str:
    for key in keys:
        value = obj.get(key)
        if isinstance(value, str) and normalize_text(value):
            return html_to_text(value)
        if isinstance(value, dict):
            nested = first_text_value(value, keys)
            if nested:
                return nested
    return ""


def author_from_value(value) -> str:
    if isinstance(value, dict):
        return normalize_text(value.get("name") or value.get("displayName") or value.get("username") or "")
    if isinstance(value, list):
        return ", ".join(filter(None, [author_from_value(item) for item in value]))
    return normalize_text(value or "")


def rating_from_object(obj: dict) -> float | None:
    rating_keys = ["ratingValue", "rating", "stars", "score", "value"]
    for key in rating_keys:
        if key not in obj:
            continue
        value = obj.get(key)
        if isinstance(value, dict):
            rating = rating_from_object(value)
        else:
            rating = rating_from_value(value)
        if rating is not None and 0 < rating <= 5:
            return rating

    for key in ["reviewRating", "starRating", "review_score", "reviewScore"]:
        value = obj.get(key)
        if isinstance(value, dict):
            rating = rating_from_object(value)
            if rating is not None and 0 < rating <= 5:
                return rating
        else:
            rating = rating_from_value(value)
            if rating is not None and 0 < rating <= 5:
                return rating
    return None


def collect_reviews_from_json(data, source: str) -> list[Review]:
    reviews: list[Review] = []
    stack = [data]
    text_keys = ["reviewBody", "reviewText", "text", "body", "content", "comment", "description", "message"]
    date_keys = ["datePublished", "publishedDate", "createdAt", "created_at", "date", "reviewDate"]

    while stack:
        item = stack.pop()
        if isinstance(item, list):
            stack.extend(item)
            continue
        if not isinstance(item, dict):
            continue

        stack.extend(value for value in item.values() if isinstance(value, (dict, list)))

        item_type = item.get("@type") or item.get("type") or item.get("__typename") or ""
        looks_like_review = bool(re.search(r"review|comment|rating", str(item_type), re.I))
        has_review_words = any(key in item for key in text_keys + ["reviewRating", "ratingValue", "stars", "score"])
        if not (looks_like_review or has_review_words):
            continue

        rating = rating_from_object(item)
        if rating is None:
            continue

        text = first_text_value(item, text_keys)
        author = author_from_value(item.get("author") or item.get("user") or item.get("customer") or item.get("reviewer"))
        date = first_text_value(item, date_keys)

        # Avoid aggregate-only ratings when there is no individual review signal.
        if not text and not author and not date and re.search(r"aggregate|summary|product", str(item_type), re.I):
            continue

        reviews.append(Review(rating=rating, text=text, author=author, date=date, source=source))
    return reviews


def extract_jsonld_reviews(html: str) -> list[Review]:
    soup = BeautifulSoup(html, "html.parser")
    reviews: list[Review] = []
    for script in soup.find_all("script", {"type": "application/ld+json"}):
        if not script.string:
            continue
        try:
            data = json.loads(script.string)
        except json.JSONDecodeError:
            continue
        reviews.extend(collect_reviews_from_json(data, "json-ld"))
    return reviews


def extract_script_json_reviews(html: str) -> list[Review]:
    soup = BeautifulSoup(html, "html.parser")
    reviews: list[Review] = []
    script_ids = {"__NEXT_DATA__", "__NUXT_DATA__", "ng-state", "apollo-state"}

    for script in soup.find_all("script"):
        raw = script.string or script.get_text("", strip=False)
        if not raw or len(raw) < 20:
            continue

        script_type = (script.get("type") or "").lower()
        script_id = script.get("id") or ""
        if script_id in script_ids or "json" in script_type:
            try:
                reviews.extend(collect_reviews_from_json(json.loads(raw), f"script-json:{script_id or script_type}"))
                continue
            except json.JSONDecodeError:
                pass

        if not re.search(r"review|rating|stars|score", raw, re.I):
            continue
        for match in re.finditer(r"(\{[^{}]{0,5000}(?:ratingValue|reviewRating|stars|score|reviewText|reviewBody)[^{}]{0,5000}\})", raw, re.I):
            try:
                reviews.extend(collect_reviews_from_json(json.loads(match.group(1)), "embedded-json-fragment"))
            except json.JSONDecodeError:
                continue
    return reviews


def star_glyph_rating(value: str) -> float | None:
    text = str(value or "")
    if not any(symbol in text for symbol in ["★", "☆", "★", "☆"]):
        return None
    filled = text.count("★")
    empty = text.count("☆")
    return float(filled) if filled and filled + empty <= 5 else None


def extract_visible_reviews(html: str) -> list[Review]:
    soup = BeautifulSoup(html, "html.parser")
    reviews: list[Review] = []
    rating_pattern = re.compile(r"([1-5](?:[.,]\d+)?)\s*(/|out of|sur|stars?|etoiles?|étoiles?|bubbles?)", re.I)
    candidates = soup.find_all(attrs={"aria-label": rating_pattern})
    candidates += soup.find_all(attrs={"title": rating_pattern})

    for node in candidates:
        label = node.get("aria-label") or node.get("title") or ""
        rating = rating_from_value(label)
        if rating is None or rating > 5:
            continue
        container = node
        for _ in range(5):
            if container.parent:
                container = container.parent
        text = normalize_text(container.get_text(" ", strip=True))
        reviews.append(Review(rating=rating, text=text[:1600], source="visible-rating-label"))

    reviewish = soup.find_all(attrs={
        "class": re.compile(r"review|avis|comment|testimonial", re.I)
    })
    reviewish += soup.find_all(attrs={
        "data-testid": re.compile(r"review|avis|comment|testimonial", re.I)
    })
    for node in reviewish:
        text = normalize_text(node.get_text(" ", strip=True))
        if len(text) < 20:
            continue
        rating = None
        for attr in ["aria-label", "title", "data-rating", "data-score"]:
            rating = rating_from_value(node.get(attr))
            if rating is not None:
                break
        if rating is None:
            rating = rating_from_value(text[:250])
        if rating is None:
            rating = star_glyph_rating(text[:250])
        if rating is None or rating > 5:
            continue
        reviews.append(Review(rating=rating, text=text[:1600], source="visible-review-block"))
    return reviews


def extract_embedded_reviews(html: str) -> list[Review]:
    reviews: list[Review] = []
    patterns = [
        r'"ratingValue"\s*:\s*"?(\d+(?:[.,]\d+)?)"?',
        r'"rating"\s*:\s*"?(\d+(?:[.,]\d+)?)"?',
        r'"stars"\s*:\s*"?(\d+(?:[.,]\d+)?)"?',
        r'"score"\s*:\s*"?(\d+(?:[.,]\d+)?)"?',
    ]
    for pattern in patterns:
        for match in re.finditer(pattern, html, flags=re.I):
            rating = rating_from_value(match.group(1))
            if rating is None or rating > 5:
                continue
            start = max(0, match.start() - 500)
            end = min(len(html), match.end() + 900)
            snippet = html_to_text(html[start:end])
            reviews.append(Review(rating=rating, text=snippet[:1600], source="embedded-rating-pattern"))
    return reviews


def extract_plain_text_reviews(text: str) -> list[Review]:
    lines = [normalize_text(line) for line in str(text or "").splitlines()]
    lines = [line for line in lines if line]
    reviews: list[Review] = []
    index = 0
    while index < len(lines):
        line = lines[index]
        rating = star_glyph_rating(line) or rating_from_value(line)
        if rating is None or rating <= 0 or rating > 5:
            index += 1
            continue

        title = lines[index + 1] if index + 1 < len(lines) else ""
        meta = lines[index + 2] if index + 2 < len(lines) else ""
        body_lines = []
        cursor = index + 3
        while cursor < len(lines):
            next_line = lines[cursor]
            if star_glyph_rating(next_line) or rating_from_value(next_line):
                break
            if re.search(r"^response from host\b|^read more\b|^see \d+ more reviews\b", next_line, re.I):
                cursor += 1
                continue
            body_lines.append(next_line)
            if len(" ".join(body_lines)) > 900:
                break
            cursor += 1
        body = normalize_text(" ".join([title] + body_lines))
        if body and len(body) > 20:
            reviews.append(Review(rating=rating, text=body[:1600], author=meta, source="visible-page-text"))
        index = max(cursor, index + 1)
    return reviews


def dedupe_reviews(reviews: Iterable[Review]) -> list[Review]:
    seen: set[str] = set()
    unique: list[Review] = []
    for review in reviews:
        if review.rating <= 0 or review.rating > 5:
            continue
        review.text = normalize_text(review.text)
        review.author = normalize_text(review.author)
        review.date = normalize_text(review.date)
        key = review.fingerprint
        if key not in seen:
            seen.add(key)
            unique.append(review)
    return unique


def fetch_reviews(product: Product) -> list[Review]:
    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(
            headless=True,
            args=[
                "--no-sandbox",
                "--disable-blink-features=AutomationControlled",
                "--disable-dev-shm-usage",
            ],
        )
        context = browser.new_context(
            locale=product.language or "en",
            timezone_id="Europe/Paris",
            viewport={"width": 1440, "height": 1200},
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/126.0.0.0 Safari/537.36"
            ),
            extra_http_headers={
                "Accept-Language": "en-US,en;q=0.9,fr;q=0.8",
                "Upgrade-Insecure-Requests": "1",
            },
        )
        context.add_init_script("""
            Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
            Object.defineProperty(navigator, 'languages', {get: () => ['en-US', 'en', 'fr']});
            Object.defineProperty(navigator, 'plugins', {get: () => [1, 2, 3, 4, 5]});
        """)
        page = context.new_page()
        page.set_default_timeout(5000)
        page.goto(product.url, wait_until="domcontentloaded", timeout=90000)
        try:
            page.wait_for_load_state("networkidle", timeout=15000)
        except Exception:
            pass
        page.wait_for_timeout(2500)

        blocked_markers = [
            "just a moment",
            "enable javascript and cookies",
            "checking your browser",
            "cf_chl",
            "challenge-platform",
        ]
        for _ in range(4):
            try:
                blocked_html = page.content().lower()
                blocked_text = page.locator("body").inner_text(timeout=5000).lower()
            except Exception:
                blocked_html = ""
                blocked_text = ""
            if not any(marker in blocked_html or marker in blocked_text for marker in blocked_markers):
                break
            page.wait_for_timeout(8000)
            try:
                page.reload(wait_until="domcontentloaded", timeout=45000)
            except Exception:
                pass

        click_labels = [
            "accept", "accepter", "j'accepte", "allow all", "tout accepter",
            "reviews", "avis", "customer reviews", "see reviews", "all reviews",
            "read reviews", "show more", "voir plus", "load more", "read more",
            "more reviews", "tous les avis", "afficher plus", "traveler reviews",
            "see all reviews", "show all reviews", "more traveler reviews"
        ]
        for label in click_labels:
            for _ in range(3):
                try:
                    page.get_by_text(re.compile(label, re.I)).first.click(timeout=1800)
                    page.wait_for_timeout(1200)
                except Exception:
                    break

        previous_height = 0
        stable_rounds = 0
        for _ in range(18):
            try:
                height = page.evaluate("document.body.scrollHeight")
            except Exception:
                height = previous_height
            page.mouse.wheel(0, 4200)
            page.wait_for_timeout(900)
            for label in ["show more", "load more", "read more", "voir plus", "afficher plus", "more reviews", "tous les avis", "traveler reviews", "see all reviews"]:
                try:
                    page.get_by_text(re.compile(label, re.I)).first.click(timeout=900)
                    page.wait_for_timeout(800)
                except Exception:
                    pass
            if height == previous_height:
                stable_rounds += 1
            else:
                stable_rounds = 0
                previous_height = height
            if stable_rounds >= 3:
                break

        html = page.content()
        try:
            visible_text = page.locator("body").inner_text(timeout=10000)
        except Exception:
            visible_text = ""
        context.close()
        browser.close()

    return dedupe_reviews([
        *extract_jsonld_reviews(html),
        *extract_script_json_reviews(html),
        *extract_visible_reviews(html),
        *extract_embedded_reviews(html),
        *extract_visible_reviews(visible_text),
        *extract_plain_text_reviews(visible_text),
    ])


def load_seen(path: Path) -> set[str]:
    if not path.exists():
        return set()
    return set(json.loads(path.read_text(encoding="utf-8")))


def save_seen(path: Path, seen: set[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(sorted(seen), ensure_ascii=False, indent=2), encoding="utf-8")


def send_email(recipients: list[str], subject: str, body: str) -> None:
    host = os.environ["SMTP_HOST"]
    port = int(os.environ.get("SMTP_PORT", "587"))
    user = os.environ.get("SMTP_USER", "")
    password = os.environ.get("SMTP_PASSWORD", "")
    sender = os.environ.get("SMTP_FROM") or user
    use_tls = os.environ.get("SMTP_USE_TLS", "true").lower() in {"true", "1", "yes", "oui"}

    message = EmailMessage()
    message["From"] = sender
    message["To"] = ", ".join(recipients)
    message["Subject"] = subject
    message.set_content(body)

    with smtplib.SMTP(host, port, timeout=30) as smtp:
        if use_tls:
            smtp.starttls()
        if user:
            smtp.login(user, password)
        smtp.send_message(message)


def try_send_email(recipients: list[str], subject: str, body: str) -> bool:
    try:
        send_email(recipients, subject, body)
        return True
    except Exception as exc:
        print(f"Email delivery failed for {', '.join(recipients)}: {exc}")
        return False


def english_subject_prefix() -> str:
    configured = os.environ.get("ALERT_SUBJECT_PREFIX", "").strip()
    if not configured:
        return "[Voxy alert]"

    lowered = configured.lower()
    if any(word in lowered for word in ["alerte", "avis"]):
        return "[Voxy alert]"
    return configured


def translate_to_english(text: str) -> str:
    if not text or os.environ.get("TRANSLATE_REVIEWS_TO_ENGLISH", "false").lower() not in {"true", "1", "yes", "oui"}:
        return ""

    api_key = os.environ.get("DEEPL_API_KEY", "").strip()
    if not api_key:
        return ""

    endpoint = "https://api-free.deepl.com/v2/translate"
    if api_key.endswith(":fx") is False:
        endpoint = "https://api.deepl.com/v2/translate"

    try:
        response = requests.post(
            endpoint,
            data={"auth_key": api_key, "text": text[:4500], "target_lang": "EN"},
            timeout=20,
        )
        response.raise_for_status()
        payload = response.json()
        return normalize_text(payload["translations"][0]["text"])
    except Exception as exc:
        return f"translation unavailable ({exc})"


def build_alert_body(product: Product, reviews: list[Review], summary: dict | None = None) -> str:
    ratings = [review.rating for review in reviews]
    avg_rating = round(sum(ratings) / len(ratings), 2) if ratings else "N/A"
    min_rating = min(ratings) if ratings else "N/A"
    critical_count = sum(1 for rating in ratings if rating < 3)
    lines = [
        "Voxy alert",
        "",
        f"Product: {product.name}",
        f"Country: {product.country or 'N/A'}",
        f"Status: {product_health_with_icon(summary)}",
        f"Trend: {trend_with_arrow(summary)}",
        f"Global score: {summary.get('global_score', 'N/A') if summary else 'N/A'}/5",
        f"New low reviews: {len(reviews)}",
        f"Average low-review score: {avg_rating}/5",
        f"Lowest review score: {min_rating}/5",
        f"Critical reviews < 3: {critical_count}",
        f"Alert rule: review <= {product.threshold}/5",
        f"Checked: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        f"Link: {product.url}",
    ]
    return "\n".join(lines)


def summarize_reviews(product: Product, reviews: list[Review]) -> dict:
    review_count = len(reviews)
    global_score = round(sum(review.rating for review in reviews) / review_count, 2) if review_count else None
    low_reviews = [review for review in reviews if review.rating <= product.threshold]
    critical_reviews = [review for review in reviews if review.rating < 3]
    all_text = " ".join(review.text.lower() for review in reviews if review.text)

    theme_counts = {}
    for theme, keywords in THEME_KEYWORDS.items():
        theme_counts[theme] = sum(all_text.count(keyword) for keyword in keywords)
    ranked_themes = [theme for theme, count in sorted(theme_counts.items(), key=lambda item: item[1], reverse=True) if count > 0]
    if not ranked_themes and review_count:
        ranked_themes = ["General satisfaction"]

    suggestions = [IMPROVEMENT_LIBRARY[theme] for theme in ranked_themes if theme in IMPROVEMENT_LIBRARY][:5]
    if not suggestions and review_count:
        suggestions = ["Read the latest low-rated reviews manually and identify recurring friction points."]

    critical_points = []
    for review in sorted(low_reviews, key=lambda item: item.rating)[:8]:
        text = review.text or "Review text not available."
        critical_points.append(f"{review.rating}/5 - {text[:260]}")

    platform = detect_platform(product)
    detailed_opportunities, source_notes = analyze_improvement_opportunities(reviews)
    return {
        "product": product.name,
        "url": product.url,
        "platform": platform,
        "country": product.country,
        "owner": product.owner,
        "priority": product.priority,
        "alert_type": product.alert_type,
        "score_threshold": product.score_threshold,
        "platform_account": product.platform_account,
        "review_count": review_count,
        "global_score": global_score,
        "trend": "No history",
        "trend_delta": "",
        "low_review_count": len(low_reviews),
        "critical_review_count": len(critical_reviews),
        "alert": bool(global_score is not None and global_score < product.score_threshold),
        "themes": ranked_themes[:6],
        "suggestions": suggestions,
        "critical_points": critical_points,
        "reviews": reviews,
        "detailed_opportunities": detailed_opportunities,
        "sources_encountered": "\n".join(source_notes),
        "status": "OK",
        "error": "",
    }


def summarize_error(product: Product, error: Exception) -> dict:
    return {
        "product": product.name,
        "url": product.url,
        "platform": detect_platform(product),
        "country": product.country,
        "owner": product.owner,
        "priority": product.priority,
        "alert_type": product.alert_type,
        "score_threshold": product.score_threshold,
        "platform_account": product.platform_account,
        "review_count": 0,
        "global_score": None,
        "trend": "No score",
        "trend_delta": "",
        "low_review_count": 0,
        "critical_review_count": 0,
        "alert": False,
        "themes": [],
        "suggestions": ["Check whether the platform blocks automation or requires a platform-specific connector."],
        "critical_points": [f"Voxy could not analyze this page: {error}"],
        "reviews": [],
        "detailed_opportunities": ["Technical issue: confirm whether the platform blocks automation, requires a platform-specific connector, or changed its page structure."],
        "sources_encountered": "No review source could be analyzed because the page check failed.",
        "status": "ERROR",
        "error": str(error),
    }


def clean_sheet_name(value: str, used: set[str]) -> str:
    base = re.sub(r"[\[\]\*:/\\?]", " ", value or "Product").strip()[:28] or "Product"
    name = base
    index = 2
    while name in used:
        suffix = f" {index}"
        name = f"{base[:31 - len(suffix)]}{suffix}"
        index += 1
    used.add(name)
    return name


def style_header(row) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in row:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def autofit_columns(sheet, min_width: int = 12, max_width: int = 70) -> None:
    for column_cells in sheet.columns:
        width = min(max_width, max(min_width, max(len(str(cell.value or "")) for cell in column_cells) + 2))
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def build_dashboard_report(summaries: list[dict], report_path: Path) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Dashboard"
    summary_sheet.append([
        "Product",
        "Country",
        "Platform",
        "Reviews detected",
        "Global score",
        "Trend",
        "Low reviews",
        "Critical reviews",
        "Alert",
        "Top themes",
        "Improvement suggestions",
        "Status",
        "Error",
        "URL",
    ])
    style_header(summary_sheet[1])

    for item in summaries:
        summary_sheet.append([
            item["product"],
            item.get("country", ""),
            item["platform"],
            item["review_count"],
            item["global_score"] if item["global_score"] is not None else "N/A",
            item.get("trend", "No history"),
            item["low_review_count"],
            item["critical_review_count"],
            "ALERT: score below 3" if item["alert"] else "OK",
            "\n".join(item["themes"]),
            "\n".join(item["suggestions"]),
            item.get("status", "OK"),
            item.get("error", ""),
            item["url"],
        ])
        if item["alert"]:
            for cell in summary_sheet[summary_sheet.max_row]:
                cell.fill = PatternFill("solid", fgColor="F8CBAD")

    used_names = {"Dashboard"}
    for item in summaries:
        sheet = workbook.create_sheet(clean_sheet_name(item["product"], used_names))
        sheet.append(["Metric", "Value"])
        style_header(sheet[1])
        metrics = [
            ("Product", item["product"]),
            ("Country", item.get("country", "")),
            ("Platform", item["platform"]),
            ("URL", item["url"]),
            ("Reviews detected", item["review_count"]),
            ("Global score", item["global_score"] if item["global_score"] is not None else "N/A"),
            ("Trend", item.get("trend", "No history")),
            ("Alert", "ALERT: score below 3" if item["alert"] else "OK"),
            ("Top themes", "\n".join(item["themes"]) or "No theme detected"),
            ("Improvement suggestions", "\n".join(item["suggestions"]) or "No suggestion available"),
            ("Critical points", "\n".join(item["critical_points"]) or "No critical point detected"),
            ("Status", item.get("status", "OK")),
            ("Error", item.get("error", "")),
        ]
        for metric, value in metrics:
            sheet.append([metric, value])

        start_row = sheet.max_row + 2
        sheet.append([])
        sheet.append(["Rating", "Author", "Date", "Source", "Review content"])
        style_header(sheet[sheet.max_row])
        for review in item["reviews"]:
            sheet.append([
                review.rating,
                review.author or "Not detected",
                review.date or "Not detected",
                review.source or "Not specified",
                review.text or "Not available on the page",
            ])
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.freeze_panes = f"A{start_row + 1}"
        autofit_columns(sheet)

    for row in summary_sheet.iter_rows(min_row=1, max_row=summary_sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    summary_sheet.freeze_panes = "A2"
    autofit_columns(summary_sheet)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(report_path)


def google_rows_for_dashboard(summaries: list[dict]) -> list[list]:
    global_summary = build_global_synthesis(summaries)
    rows = [
        ["Voxy daily alert summary", ""],
        ["Products checked", global_summary["product_count"]],
        ["Average score", global_summary["average_score"] if global_summary["average_score"] is not None else "N/A"],
        ["Low reviews", global_summary["low_reviews"]],
        ["Critical reviews", global_summary["critical_reviews"]],
        [],
        [
        "Product",
        "Country",
        "Owner",
        "Priority",
        "Status",
        "Trend",
        "Global score",
        "Low reviews",
        "Critical reviews",
        "Main issue",
        "Action",
        "URL",
        ],
    ]
    for item in summaries:
        rows.append([
            item["product"],
            item.get("country", ""),
            item.get("owner", ""),
            item.get("priority", ""),
            product_health_with_icon(item),
            trend_with_arrow(item),
            item["global_score"] if item["global_score"] is not None else "N/A",
            item["low_review_count"],
            item["critical_review_count"],
            ", ".join(item.get("themes", [])[:2]) or "No issue detected",
            (item.get("suggestions") or ["No action needed"])[0],
            item["url"],
        ])
    return rows


def google_rows_for_product(summary: dict) -> list[list]:
    rows = [
        ["Metric", "Value"],
        ["Product", summary["product"]],
        ["Country", summary.get("country", "")],
        ["Platform", summary["platform"]],
        ["URL", summary["url"]],
        ["Reviews detected", summary["review_count"]],
        ["Global score", summary["global_score"] if summary["global_score"] is not None else "N/A"],
        ["Trend", summary.get("trend", "No history")],
        ["Alert", "ALERT: score below 3" if summary["alert"] else "OK"],
        ["Top themes", "\n".join(summary["themes"]) or "No theme detected"],
        ["Improvement suggestions", "\n".join(summary["suggestions"]) or "No suggestion available"],
        ["Detailed improvement opportunities", "\n\n".join(summary.get("detailed_opportunities", [])) or "No detailed opportunity available"],
        ["Critical points", "\n".join(summary["critical_points"]) or "No critical point detected"],
        ["Sources encountered", summary.get("sources_encountered", "No source detected")],
        ["Status", summary.get("status", "OK")],
        ["Error", summary.get("error", "")],
        [],
        ["Rating", "Author", "Date", "Source", "Review content"],
    ]
    for review in summary["reviews"]:
        rows.append([
            review.rating,
            clean_author_for_email(review),
            clean_date_for_email(review),
            review.source or "Not specified",
            clean_review_text_for_email(review.text, max_chars=900) or "Not available on the page",
        ])
    return rows


def rectangularize_rows(rows: list[list]) -> list[list]:
    if not rows:
        return rows
    max_cols = max(len(row) for row in rows)
    return [row + [""] * (max_cols - len(row)) for row in rows]


def parse_optional_float(value) -> float | None:
    if value is None:
        return None
    text = str(value).strip().replace(",", ".")
    if not text or text.upper() in {"N/A", "NA", "NONE"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def trend_label(current_score: float | None, previous_score: float | None, sensitivity: float = 0.10) -> tuple[str, str]:
    if current_score is None:
        return "No score", ""
    if previous_score is None:
        return "No history", ""
    delta = round(current_score - previous_score, 2)
    if delta >= sensitivity:
        return f"Improving (+{delta:.2f})", f"+{delta:.2f}"
    if delta <= -sensitivity:
        return f"Declining ({delta:.2f})", f"{delta:.2f}"
    sign = "+" if delta > 0 else ""
    return f"Stable ({sign}{delta:.2f})", f"{sign}{delta:.2f}"


def latest_scores_from_history(spreadsheet) -> dict[str, float]:
    try:
        history = spreadsheet.worksheet("History")
    except gspread.WorksheetNotFound:
        return {}
    rows = history.get_all_values()
    if len(rows) < 2:
        return {}
    headers = [header.strip().lower() for header in rows[0]]
    try:
        product_index = headers.index("product")
        score_index = headers.index("global score")
    except ValueError:
        return {}

    latest: dict[str, float] = {}
    for row in rows[1:]:
        if len(row) <= max(product_index, score_index):
            continue
        product = row[product_index].strip()
        score = parse_optional_float(row[score_index])
        if product and score is not None:
            latest[product] = score
    return latest


def annotate_trends_from_history(spreadsheet, summaries: list[dict]) -> None:
    previous_scores = latest_scores_from_history(spreadsheet)
    for summary in summaries:
        label, delta = trend_label(summary.get("global_score"), previous_scores.get(summary["product"]))
        summary["trend"] = label
        summary["trend_delta"] = delta


def append_history_rows(spreadsheet, summaries: list[dict]) -> None:
    history = get_or_create_worksheet(spreadsheet, "History", rows=max(100, len(summaries) + 20), cols=12)
    rows = history.get_all_values()
    headers_with_country = [
        "Run timestamp",
        "Product",
        "Country",
        "Platform",
        "Reviews detected",
        "Global score",
        "Trend",
        "Low reviews",
        "Critical reviews",
        "Status",
        "URL",
    ]
    if not rows:
        history.update([headers_with_country], value_input_option="USER_ENTERED")
        use_country_column = True
    else:
        current_headers = [header.strip().lower() for header in rows[0]]
        use_country_column = "country" in current_headers

    run_timestamp = datetime.now(ZoneInfo("Europe/Paris")).strftime("%Y-%m-%d %H:%M")
    new_rows = []
    for summary in summaries:
        row = [
            run_timestamp,
            summary["product"],
            summary["platform"],
            summary["review_count"],
            summary["global_score"] if summary["global_score"] is not None else "N/A",
            summary.get("trend", "No history"),
            summary["low_review_count"],
            summary["critical_review_count"],
            summary.get("status", "OK"),
            summary["url"],
        ]
        if use_country_column:
            row.insert(2, summary.get("country", ""))
        new_rows.append(row)
    if new_rows:
        history.append_rows(new_rows, value_input_option="USER_ENTERED")


def get_or_create_worksheet(spreadsheet, title: str, rows: int = 100, cols: int = 12):
    try:
        return spreadsheet.worksheet(title)
    except gspread.WorksheetNotFound:
        return spreadsheet.add_worksheet(title=title, rows=rows, cols=cols)


def update_google_sheet_dashboard(sheet_url: str, summaries: list[dict]) -> None:
    client = google_client_from_env()
    if client is None:
        raise RuntimeError("GOOGLE_SERVICE_ACCOUNT_JSON is required to update the shared Google Sheet dashboard.")
    spreadsheet = client.open_by_url(sheet_url)
    annotate_trends_from_history(spreadsheet, summaries)

    dashboard = get_or_create_worksheet(spreadsheet, "Dashboard", rows=max(100, len(summaries) + 10), cols=11)
    dashboard.clear()
    dashboard.update(rectangularize_rows(google_rows_for_dashboard(summaries)), value_input_option="USER_ENTERED")
    dashboard.freeze(rows=1)

    append_history_rows(spreadsheet, summaries)
    print("Shared Google Sheet alert summary updated.")


def build_score_alert_body(summary: dict) -> str:
    return "\n".join([
        "Voxy alert",
        "",
        f"Product: {summary['product']}",
        f"Country: {summary.get('country') or 'N/A'}",
        f"Status: {product_health_with_icon(summary)}",
        f"Trend: {trend_with_arrow(summary)}",
        f"Global score: {summary['global_score']}/5",
        f"Reviews detected: {summary['review_count']}",
        f"Low reviews: {summary['low_review_count']}",
        f"Critical reviews < 3: {summary['critical_review_count']}",
        f"Main issue: {', '.join(summary['themes'][:2]) or 'not detected'}",
        f"Link: {summary['url']}",
    ])


def alert_type_allows(product: Product, alert_kind: str) -> bool:
    value = (product.alert_type or "Both").strip().lower()
    if value in {"both", "all", "tout"}:
        return True
    if alert_kind == "low_review":
        return value in {"low reviews", "low review", "review", "reviews", "avis", "mauvais avis"}
    if alert_kind == "score":
        return value in {"score", "score drop", "global score", "note"}
    return False


def build_summary_email_body(recipient: str, entries: list[dict], timezone_name: str) -> str:
    now = datetime.now(ZoneInfo(timezone_name)).strftime("%Y-%m-%d %H:%M")
    lines = [
        "Voxy weekly review summary",
        "",
        f"Recipient: {recipient}",
        f"Check time: {now} ({timezone_name})",
        f"Products checked in this email: {len(entries)}",
        "",
    ]
    for index, entry in enumerate(entries, start=1):
        summary = entry["summary"]
        reasons = ", ".join(entry["reasons"])
        lines.extend([
            f"{index}. {summary['product']}",
            f"Country: {summary.get('country') or 'N/A'}",
            f"Owner: {summary.get('owner') or 'N/A'}",
            f"Priority: {summary.get('priority') or 'Medium'}",
            f"Platform: {summary.get('platform') or 'N/A'}",
            f"Status: {product_health_with_icon(summary)}",
            f"Trend: {trend_with_arrow(summary)}",
            f"Score: {summary['global_score'] if summary['global_score'] is not None else 'N/A'}/5",
            f"Score alert threshold: {summary.get('score_threshold', 3.0)}/5",
            f"Reviews checked: {summary['review_count']}",
            f"Low reviews: {summary['low_review_count']}",
            f"Critical reviews < 3: {summary['critical_review_count']}",
            f"Reason: {reasons}",
            f"Action: {(summary.get('suggestions') or ['Review product feedback'])[0]}",
            f"Link: {summary['url']}",
            "",
        ])
    return "\n".join(lines).strip()


def seen_key(product: Product, review: Review) -> str:
    product_key = hashlib.sha256(product.url.encode("utf-8", errors="ignore")).hexdigest()[:16]
    return f"{product_key}:{review.fingerprint}"


def score_alert_key(product: Product, summary: dict) -> str:
    product_key = hashlib.sha256(product.url.encode("utf-8", errors="ignore")).hexdigest()[:16]
    score = summary["global_score"] if summary["global_score"] is not None else "na"
    return f"{product_key}:score-alert:{score}:{summary['review_count']}"


def annotate_email_trends(sheet_url: str, summaries: list[dict]) -> None:
    if not sheet_url or not summaries:
        return
    client = google_client_from_env()
    if client is None:
        return
    try:
        spreadsheet = client.open_by_url(sheet_url)
        annotate_trends_from_history(spreadsheet, summaries)
    except Exception as exc:
        print(f"Trend check skipped before email: {exc}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Monitor reviews across multiple platforms and send email alerts.")
    parser.add_argument("--xlsx", default="template_surveillance_avis_multi_plateformes.xlsx", help="Chemin du fichier Excel.")
    parser.add_argument("--sheet-url", default="", help="Google Sheets URL to download before running.")
    parser.add_argument("--state", default="avis_deja_signales.json", help="Memory file for reviews already reported.")
    parser.add_argument("--report", default="voxy_dashboard_report.xlsx", help="Dashboard report Excel file to generate.")
    parser.add_argument("--timezone", default="Europe/Paris", help="Timezone used to decide whether a review date is from today.")
    parser.add_argument("--only-at-hour", type=int, default=None, help="Exit unless the current hour in --timezone matches this value.")
    parser.add_argument("--update-google-sheet-dashboard", action="store_true", help="Write Dashboard and product tabs back into the shared Google Sheet.")
    parser.add_argument("--include-past-dated-reviews", action="store_true", help="Include reviews whose detected date is before today.")
    parser.add_argument("--dry-run", action="store_true", help="Test without sending email.")
    parser.add_argument("--baseline", action="store_true", help="Save current reviews without sending alerts.")
    args = parser.parse_args()

    load_dotenv()
    sheet_url = args.sheet_url or os.environ.get("GOOGLE_SHEET_URL", "")
    input_xlsx = Path(args.xlsx)
    if sheet_url and os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip():
        products = load_products_from_google_sheet(sheet_url)
    elif sheet_url:
        input_xlsx = download_google_sheet(sheet_url, Path("voxy_current_shared_sheet.xlsx"))
        print(f"Downloaded shared Google Sheet to: {input_xlsx}")
        products = load_products(input_xlsx)
    else:
        products = load_products(input_xlsx)
    seen = load_seen(Path(args.state))
    new_seen = set(seen)
    subject_prefix = english_subject_prefix()
    local_now = datetime.now(ZoneInfo(args.timezone))
    local_today = local_now.date()

    if args.only_at_hour is not None and local_now.hour != args.only_at_hour:
        print(f"Not the scheduled local hour yet: {local_now.strftime('%Y-%m-%d %H:%M')} ({args.timezone}).")
        return 0

    if not products:
        print("No active products found in the Excel file.")
        return 0

    summaries: list[dict] = []
    weekly_entries_by_recipient: dict[str, list[dict]] = {}
    for product in products:
        print(f"Checking: {product.name}")
        try:
            reviews = fetch_reviews(product)
        except Exception as exc:
            print(f"Error while checking {product.name}: {exc}")
            summary = summarize_error(product, exc)
            summaries.append(summary)
            if not args.baseline:
                for recipient in product.emails:
                    weekly_entries_by_recipient.setdefault(recipient, []).append({
                        "summary": summary,
                        "reasons": ["technical check needed"],
                    })
            continue

        for review in reviews:
            new_seen.add(seen_key(product, review))

        if args.include_past_dated_reviews:
            current_reviews = reviews
        else:
            current_reviews = [review for review in reviews if not is_past_dated_review(review, local_today)]
            skipped = len(reviews) - len(current_reviews)
            if skipped:
                print(f"Skipped {skipped} review(s) dated before {local_today.isoformat()} ({args.timezone}).")

        summary = summarize_reviews(product, current_reviews)
        summaries.append(summary)

        low_reviews = [review for review in current_reviews if review.rating <= product.threshold and seen_key(product, review) not in seen]

        if args.baseline:
            print(f"Baseline: {len(reviews)} reviews saved, no email sent.")
            continue

        alert_reasons = []
        if summary["alert"] and score_alert_key(product, summary) not in seen and alert_type_allows(product, "score"):
            alert_reasons.append(f"score below {product.score_threshold}")
            new_seen.add(score_alert_key(product, summary))

        if not low_reviews:
            print(f"OK: no new reviews rated {product.threshold} stars or less.")
        else:
            if alert_type_allows(product, "low_review"):
                alert_reasons.append(f"{len(low_reviews)} new review(s) <= {product.threshold}/5")

        if alert_reasons:
            for recipient in product.emails:
                weekly_entries_by_recipient.setdefault(recipient, []).append({
                    "summary": summary,
                    "reasons": alert_reasons,
                })
        elif not args.baseline:
            for recipient in product.emails:
                weekly_entries_by_recipient.setdefault(recipient, []).append({
                    "summary": summary,
                    "reasons": ["weekly status: no new alert"],
                })

    if weekly_entries_by_recipient and not args.baseline:
        annotate_email_trends(sheet_url, summaries)
        for recipient, entries in sorted(weekly_entries_by_recipient.items()):
            body = build_summary_email_body(recipient, entries, args.timezone)
            alert_count = sum(
                1 for entry in entries
                if "weekly status: no new alert" not in entry["reasons"]
            )
            subject = f"{subject_prefix} Voxy weekly summary: {len(entries)} product(s) checked, {alert_count} in alert"
            if args.dry_run:
                print("DRY RUN - weekly summary email not sent")
                print(body)
            else:
                if try_send_email([recipient], subject, body):
                    print(f"Weekly summary sent to: {recipient}")
    elif not args.baseline:
        print("OK: no weekly summary recipients found.")

    if summaries:
        build_dashboard_report(summaries, Path(args.report))
        print(f"Dashboard report saved: {args.report}")
        if args.update_google_sheet_dashboard:
            if not sheet_url:
                raise RuntimeError("--update-google-sheet-dashboard requires GOOGLE_SHEET_URL or --sheet-url.")
            update_google_sheet_dashboard(sheet_url, summaries)

    save_seen(Path(args.state), new_seen)
    return 0


if __name__ == "__main__":
    sys.exit(main())
